from openpyxl import load_workbook, Workbook
import warnings
import time

nombre_modelo = str(input("Ingrese el nombre del modelo: "))
nombre1_pago = str(input("Ingrese el nombre de la planilla de pagos: "))

nombre_modelo = f'{nombre_modelo}.xlsx'
nombre_pago = f'{nombre1_pago}.xlsx'

modelo = load_workbook(nombre_modelo, data_only=True)
pago = load_workbook(nombre_pago, data_only=True)

ws_modelo = modelo[modelo.sheetnames[0]]
ws_pago = pago[pago.sheetnames[0]]

dnis = []
montos = []
dni_montos = {}
for column_data in ws_modelo['C']:
    if column_data.value != None and column_data.value != 'DNI':
        dnis.append(str(column_data.value))

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    for column_data in ws_modelo['AG']:
        if column_data.value != None and column_data.value != 'BANCO':
            montos.append(str(column_data.value))
        if column_data.value == None:
            montos.append('0')

montos_final = []
for monto in montos:
    #print(round(float(monto), 2))
    if monto != ' ':
        monto = float(monto)
        monto = f'{monto:.2f}'
    else:
        monto = 0,00
    montos_final.append(monto)


contador = 0
print(len(dnis))
print(len(montos_final))
for dni in dnis:
    montos_final[contador] = montos_final[contador].replace('.', ',')
    dni_montos[dni] = montos_final[contador]
    contador += 1 


for row in ws_pago.iter_rows():
    
    if row[2].value in dnis:
        row[5].value = dni_montos[row[2].value]

pago.save(f"{nombre1_pago}-INTEGRADO.xlsx")

print('Completado con éxito.')

while True:
    time.sleep(1)
    print('Completado con éxito. Puede cerrar el Script.')